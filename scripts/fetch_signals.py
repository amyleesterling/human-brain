"""Pull real recordings of real human brains, for the time axis of the scale map.

H01 is one instant fixed in resin and the tractography is a population average
of shapes. Neither has a time axis at all. Everything on the scale map that
happens rather than merely exists needs a recording, and these are recordings
of people.

  action potential  DANDI 000293. Whole cell current clamp from a cortical
                    neuron of a 59 year old human. CC0-1.0.
  eeg               PhysioNet EEG Motor Movement/Imagery Database, subject 1,
                    runs 1 and 2, eyes open and eyes closed. ODC-BY 1.0.
  seeg              OpenNeuro ds007095. Human hippocampal depth electrodes.
                    CC0.

Every number the page prints about these is measured here, printed here, and
written into data/signals.json beside the trace it came from. Licences differ
per file and travel with the data.

Usage: python scripts/fetch_signals.py
"""

import json
import os
import struct
import urllib.request

import numpy as np

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CACHE = os.path.join(ROOT, ".cache")

AP_URL = ("https://api.dandiarchive.org/api/dandisets/000293/versions/"
          "0.220708.1652/assets/747e0b3a-b61b-45c5-94cd-f2cd9b8e80cd/download/")
EEG = {
    "closed": "https://physionet.org/files/eegmmidb/1.0.0/S001/S001R02.edf",
    "open": "https://physionet.org/files/eegmmidb/1.0.0/S001/S001R01.edf",
}
SEEG_URL = ("https://s3.amazonaws.com/openneuro.org/ds007095/sub-01/ses-01/"
            "ieeg/sub-01_ses-01_task-seizure_run-01_ieeg.edf")
MYELIN_URL = ("https://static-content.springer.com/esm/"
              "art%3A10.1038%2Fs41598-018-33112-8/MediaObjects/"
              "41598_2018_33112_MOESM2_ESM.xlsx")


def p(*a):
    q = os.path.join(ROOT, *a)
    os.makedirs(os.path.dirname(q), exist_ok=True)
    return q


def cached(url, name):
    os.makedirs(CACHE, exist_ok=True)
    dst = os.path.join(CACHE, name)
    if not os.path.exists(dst):
        print(f"  downloading {name}")
        urllib.request.urlretrieve(url, dst)
    return dst


# ------------------------------------------------------------------ EDF ----
def clean_label(s):
    """EDF pads channel names to 16 characters. This dataset pads with dots,
    so the occipital midline channel arrives as "Oz.." and prints that way."""
    return s.strip().strip(".").strip()


def read_edf(path):
    """EDF is a 256 byte header, one 256 byte block per signal, then records of
    int16. Simple enough to read directly rather than take a dependency."""
    with open(path, "rb") as f:
        raw = f.read()
    ns = int(raw[252:256])
    nrec = int(raw[236:244])
    dur = float(raw[244:252])
    o = 256
    def block(n, w):
        nonlocal o
        v = [raw[o + i * w: o + (i + 1) * w].decode("ascii", "replace").strip()
             for i in range(n)]
        o += n * w
        return v
    labels = block(ns, 16)
    block(ns, 80)                      # transducer
    units = block(ns, 8)
    pmin = np.array(block(ns, 8), float)
    pmax = np.array(block(ns, 8), float)
    dmin = np.array(block(ns, 8), float)
    dmax = np.array(block(ns, 8), float)
    block(ns, 80)                      # prefiltering
    spr = np.array(block(ns, 8), int)  # samples per record per signal
    o += ns * 32
    data = np.frombuffer(raw, "<i2", offset=o,
                         count=nrec * int(spr.sum())).reshape(nrec, int(spr.sum()))
    out, start = {}, 0
    gain = (pmax - pmin) / (dmax - dmin)
    for i, lab in enumerate(labels):
        seg = data[:, start:start + spr[i]].reshape(-1).astype(np.float64)
        out[lab] = seg * gain[i] + (pmin[i] - dmin[i] * gain[i])
        start += spr[i]
    fs = {lab: spr[i] / dur for i, lab in enumerate(labels)}
    return out, fs, units, nrec * dur


def bandpower(x, fs, lo, hi):
    f = np.fft.rfftfreq(len(x), 1 / fs)
    pw = np.abs(np.fft.rfft(x * np.hanning(len(x)))) ** 2
    m = (f >= lo) & (f < hi)
    return float(pw[m].sum()), f, pw


# ---------------------------------------------------- action potential ----
def get_ap():
    import h5py

    f = h5py.File(cached(AP_URL, "dandi000293_cell.nwb"), "r")
    acq = f["acquisition"]
    sweeps = sorted(k for k in acq if "data" in acq[k])
    best, bestn = None, -1
    for k in sweeps:
        g = acq[k]
        if "data" not in g:
            continue
        d = np.asarray(g["data"][()], float)
        if d.ndim != 1 or len(d) < 100:
            continue
        rate = float(g["starting_time"].attrs.get("rate", 0)) if "starting_time" in g else 0
        if rate <= 0:
            continue
        # The NWB conversion attribute on this file says 1e-12 but the values
        # are already millivolts. Trust the range, not the attribute: a trace
        # peaking near +14 with a baseline near -70 is mV and nothing else.
        if d.max() > 500 or d.min() < -500:
            continue
        n = int((np.diff((d > 0).astype(int)) > 0).sum())
        if n > bestn:
            best, bestn, brate = k, n, rate
    d = np.asarray(acq[best]["data"][()], float)
    fs = brate
    print(f"  sweep {best}: {len(d)} samples at {fs:.0f} Hz, "
          f"{d.min():.1f} to {d.max():.1f} mV, {bestn} upward zero crossings")

    # Take the first spike and cut a window around its peak.
    over = np.where(d > 0)[0]
    pk = over[0] + int(np.argmax(d[over[0]:over[0] + int(0.005 * fs)]))
    pre, post = int(0.002 * fs), int(0.004 * fs)
    seg = d[pk - pre: pk + post]
    t = (np.arange(len(seg)) - pre) / fs * 1000.0

    # Volts per second. seg is millivolts and the sample spacing is 1/fs
    # seconds, so gradient gives mV/s and the thousand converts it. Getting
    # this wrong is silent: the first version divided by fs and by a thousand
    # and reported an upstroke of 300 million volts per second.
    dv = np.gradient(seg, 1.0 / fs) / 1000.0
    # Threshold is conventionally where the rate of rise passes 20 V/s.
    thr_i = int(np.argmax(dv > 20))
    peak = float(seg.max())
    thr = float(seg[thr_i])
    half = thr + (peak - thr) / 2
    above = np.where(seg >= half)[0]
    hw = float((above[-1] - above[0]) / fs * 1000.0)
    stats = {
        "threshold_mV": round(thr, 1), "peak_mV": round(peak, 1),
        "amplitude_mV": round(peak - thr, 1), "half_width_ms": round(hw, 3),
        "max_upstroke_V_per_s": round(float(dv.max()), 0),
        "max_downstroke_V_per_s": round(float(dv.min()), 0),
        "sample_rate_Hz": round(fs),
    }
    for k, v in stats.items():
        print(f"    {k:24s} {v}")
    # A human cortical action potential is tens of millivolts tall, under a
    # millisecond wide, and rises at hundreds of volts per second. Anything
    # outside that is a units mistake, not a discovery.
    assert 20 < stats["amplitude_mV"] < 150, stats
    assert 0.1 < stats["half_width_ms"] < 3.0, stats
    assert 20 < stats["max_upstroke_V_per_s"] < 1500, stats
    return {
        "id": "action-potential",
        "title": "One human neuron firing",
        "source": AP_URL, "dandiset": "DANDI:000293", "licence": "CC0-1.0",
        "citation": "Whole cell current clamp from human cortical tissue, "
                    "DANDI dandiset 000293.",
        "subject": "human cortical neuron",
        "x_label": "milliseconds", "y_label": "millivolts",
        "x": [round(float(v), 4) for v in t],
        "y": [round(float(v), 2) for v in seg],
        "stats": stats,
    }


# ---------------------------------------------------------------- EEG ----
def get_eeg():
    out = {}
    for state, url in EEG.items():
        sig, fs, units, dur = read_edf(cached(url, f"eegmmidb_S001_{state}.edf"))
        key = next(k for k in sig if k.replace(".", "").strip().lower() == "oz")
        x = sig[key]
        f0 = fs[key]
        seg = x[int(20 * f0): int(24 * f0)]
        alpha, _, _ = bandpower(x[int(5 * f0):int(55 * f0)], f0, 8, 13)
        theta, _, _ = bandpower(x[int(5 * f0):int(55 * f0)], f0, 4, 8)
        beta, ff, pw = bandpower(x[int(5 * f0):int(55 * f0)], f0, 13, 30)
        band = (ff >= 6) & (ff <= 14)
        peak = float(ff[band][np.argmax(pw[band])])
        out[state] = {
            "state": state, "channel": clean_label(key), "fs": f0,
            "alpha_over_theta": round(alpha / theta, 2),
            "alpha_over_beta": round(alpha / beta, 2),
            "peak_Hz": round(peak, 2),
            "x": [round(i / f0, 4) for i in range(len(seg))],
            "y": [round(float(v), 1) for v in seg],
            "source": url,
        }
        print(f"  eyes {state:6s} {key.strip()}: alpha/theta "
              f"{out[state]['alpha_over_theta']:6.2f}  peak "
              f"{out[state]['peak_Hz']:.2f} Hz")
    ratio = out["closed"]["alpha_over_theta"] / out["open"]["alpha_over_theta"]
    print(f"  alpha blocking: closed is {ratio:.1f} times open")
    return {
        "id": "eeg-alpha",
        "title": "A rhythm across the back of one head",
        "licence": "ODC-BY 1.0",
        "citation": "PhysioNet EEG Motor Movement/Imagery Database, subject 1, "
                    "runs 1 and 2. Goldberger et al., Circulation 101:e215 (2000); "
                    "Schalk et al., IEEE TBME 51:1034 (2004).",
        "x_label": "seconds", "y_label": "microvolts",
        "blocking_ratio": round(ratio, 1),
        "traces": out,
    }


# --------------------------------------------------------------- sEEG ----
def get_seeg():
    sig, fs, units, dur = read_edf(cached(SEEG_URL, "ds007095_seeg.edf"))
    print(f"  {len(sig)} channels, {dur:.0f} s")
    picked, report = None, []
    for lab, x in sig.items():
        f0 = fs[lab]
        # A channel that never goes positive and piles up on one value is
        # clipped at the amplifier rail and must not be plotted as signal.
        rail = float(np.mean(np.isclose(x, x.min(), atol=1e-6)))
        report.append((lab, float(x.min()), float(x.max()), rail))
        print(f"    {lab.strip():18s} {x.min():9.1f} to {x.max():8.1f} "
              f"{units[list(sig).index(lab)]:4s}  {rail*100:5.2f}% at the floor")
        if x.max() > 0 and rail < 0.001 and picked is None:
            picked, pf = lab, f0
    if picked is None:
        raise SystemExit("every channel looks clipped, refusing to plot one")
    x = sig[picked]
    seg = x[int(10 * pf): int(14 * pf)]
    print(f"  plotting {picked.strip()}, the one channel that is not clipped")
    return {
        "id": "seeg",
        "title": "Inside a living human hippocampus",
        "source": SEEG_URL, "licence": "CC0",
        "citation": "OpenNeuro ds007095, human stereo EEG.",
        "channel": clean_label(picked), "fs": pf,
        "x_label": "seconds", "y_label": "microvolts",
        "x": [round(i / pf, 4) for i in range(len(seg))],
        "y": [round(float(v), 1) for v in seg],
        "channel_report": [
            {"channel": a.strip(), "min": round(b, 1), "max": round(c, 1),
             "fraction_at_floor": round(d, 4)} for a, b, c, d in report],
    }


# -------------------------------------------------------------- bands ----
BANDS = [
    ("delta", 1.0, 4.0, "Deep sleep, and the slowest thing a scalp can hear."),
    ("theta", 4.0, 8.0, "Drowsiness and navigation. Sets the windows the faster "
                        "rhythms are allowed to matter in."),
    ("alpha", 8.0, 13.0, "The idling rhythm of the back of the head. Present "
                         "with the eyes closed, gone when they open."),
    ("beta", 13.0, 30.0, "Alert, engaged, holding still."),
    ("gamma", 30.0, 45.0, "Fast enough that a single spike can be early or late "
                          "within one cycle."),
]


def get_bands():
    """The same one recording, split into the five named rhythms.

    Everything here is one channel of one person: nothing is simulated and no
    band is drawn from a different source than any other. The bands are the
    conventional clinical ranges and the power in each is measured, so the
    reason alpha towers over the rest with the eyes closed is the recording's,
    not a choice made while drawing it.
    """
    from scipy.signal import butter, sosfiltfilt, welch

    out = {}
    for state, url in EEG.items():
        sig, fs, units, dur = read_edf(cached(url, f"eegmmidb_S001_{state}.edf"))
        key = next(k for k in sig if k.replace(".", "").strip().lower() == "oz")
        x = sig[key]
        f0 = fs[key]
        # Analyse a clean stretch away from both ends of the run.
        seg_full = x[int(5 * f0):int(55 * f0)]
        ff, pw = welch(seg_full, f0, nperseg=int(4 * f0))

        # 4 seconds is long enough to show four cycles of delta and hundreds of
        # gamma, which is what makes the pile legible.
        t0, t1 = int(20 * f0), int(24 * f0)
        traces, power = {}, {}
        for name, lo, hi in [(b[0], b[1], b[2]) for b in BANDS]:
            # Butterworth as second order sections, applied forwards and
            # backwards so the filter adds no phase shift. A one-directional
            # filter would slide each band in time against the others and the
            # stack would be lying about when things happen.
            sos = butter(4, [lo / (f0 / 2), hi / (f0 / 2)], btype="band",
                         output="sos")
            y = sosfiltfilt(sos, seg_full)
            traces[name] = [round(float(v), 2)
                            for v in y[t0 - int(5 * f0):t1 - int(5 * f0)]]
            m = (ff >= lo) & (ff < hi)
            power[name] = float(np.trapezoid(pw[m], ff[m]))

        tot = sum(power.values())
        out[state] = {
            "state": state, "channel": clean_label(key), "fs": f0,
            "seconds": 4.0,
            "traces": traces,
            "power": {k: round(v, 2) for k, v in power.items()},
            "share": {k: round(100 * v / tot, 1) for k, v in power.items()},
            "spectrum": {
                "f": [round(float(v), 2) for v in ff[(ff >= 1) & (ff <= 45)]],
                "p": [round(float(v), 3) for v in pw[(ff >= 1) & (ff <= 45)]],
            },
            "source": url,
        }
        sh = out[state]["share"]
        print(f"  eyes {state:6s} " +
              "  ".join(f"{k} {sh[k]:5.1f}%" for k in sh))

    ao, ac = out["open"]["share"]["alpha"], out["closed"]["share"]["alpha"]
    print(f"  alpha share of total power: closed {ac}%, open {ao}%, "
          f"ratio {ac/ao:.1f}")
    # The whole point of showing five bands from one recording is that closing
    # the eyes moves power into exactly one of them.
    assert ac > ao, (ac, ao)
    return {
        "id": "eeg-bands",
        "title": "The five rhythms, from one recording",
        "licence": "ODC-BY 1.0",
        "citation": "PhysioNet EEG Motor Movement/Imagery Database, subject 1, "
                    "runs 1 and 2. Goldberger et al., Circulation 101:e215 "
                    "(2000); Schalk et al., IEEE TBME 51:1034 (2004).",
        "x_label": "seconds", "y_label": "microvolts",
        "method": "Fourth order Butterworth band pass, applied forwards and "
                  "backwards so no band is shifted in time against another. "
                  "Power is the integral of Welch's spectrum over each band.",
        "bands": [{"name": n, "low_Hz": lo, "high_Hz": hi, "note": d}
                  for n, lo, hi, d in BANDS],
        "alpha_share_ratio": round(ac / ao, 1),
        "states": out,
    }


# ------------------------------------------------------------ topography ----
def get_topography():
    """Where on the head each rhythm actually is, from all 64 electrodes.

    The recording is a full 10-10 montage, not one channel, so the question
    "do the different waves come from different places" is one this file can
    answer rather than assert. Electrode positions are the standard 10-05
    coordinates, which is a template head and not this person's, so what is
    real here is the pattern across sites and not the millimetres.
    """
    import mne
    from scipy.signal import welch

    mon = mne.channels.make_standard_montage("standard_1005")
    pos3 = {k.lower(): v for k, v in mon.get_positions()["ch_pos"].items()}

    states = {}
    chans = None
    for state, url in EEG.items():
        sig, fs, units, dur = read_edf(cached(url, f"eegmmidb_S001_{state}.edf"))
        rows, missing = [], []
        for k in sig:
            if "annotation" in k.lower():
                continue
            lab = clean_label(k)
            p3 = pos3.get(lab.lower())
            if p3 is None:
                missing.append(lab)
                continue
            f0 = fs[k]
            x = sig[k][int(5 * f0):int(55 * f0)]
            ff, pw = welch(x, f0, nperseg=int(4 * f0))
            band = (ff >= 1) & (ff <= 45)
            tot = float(np.trapezoid(pw[band], ff[band]))
            share = {}
            for name, lo, hi, _ in BANDS:
                m = (ff >= lo) & (ff < hi)
                share[name] = round(
                    100 * float(np.trapezoid(pw[m], ff[m])) / tot, 2)
            rows.append({"ch": lab, "xyz_mm": [round(float(v) * 1000, 1) for v in p3],
                         "share": share})
        if missing:
            raise SystemExit(f"no standard position for {missing}")
        rows.sort(key=lambda r: r["ch"])
        if chans is None:
            chans = [r["ch"] for r in rows]
        else:
            assert chans == [r["ch"] for r in rows], "montage differs between runs"
        states[state] = rows
        print(f"  eyes {state:6s} {len(rows)} electrodes placed")

    # The classic round head view is an azimuthal projection seen from above:
    # distance from the centre is the angle down from the vertex, and the angle
    # around is the angle around the head. Electrodes below the equator land
    # outside the drawn circle, which is where they physically are.
    P = np.array([r["xyz_mm"] for r in states["closed"]], float)
    centre = np.array([0.0, 0.0, 0.0])
    v = P - centre
    r = np.linalg.norm(v, axis=1)
    theta = np.arccos(np.clip(v[:, 2] / r, -1, 1))     # down from +z
    phi = np.arctan2(v[:, 1], v[:, 0])                  # around, +y is the nose
    rad = theta / (np.pi / 2)
    xy = np.column_stack([rad * np.cos(phi), rad * np.sin(phi)])
    for i, row in enumerate(states["closed"]):
        row["xy"] = [round(float(xy[i, 0]), 4), round(float(xy[i, 1]), 4)]
    by = {r["ch"]: r["xy"] for r in states["closed"]}
    for row in states["open"]:
        row["xy"] = by[row["ch"]]

    # Where each band is strongest, measured rather than remembered.
    peaks = {}
    for name, lo, hi, _ in BANDS:
        best = max(states["closed"], key=lambda r: r["share"][name])
        peaks[name] = {"channel": best["ch"], "share": best["share"][name]}
        print(f"    {name:6s} peaks at {best['ch']:5s} "
              f"{best['share'][name]:5.1f}% of that site's power")
    # Alpha over the back of the head is the one result this has to reproduce.
    assert peaks["alpha"]["channel"].lower() in {
        "oz", "o1", "o2", "poz", "po3", "po4", "po7", "po8", "pz", "p8", "p7"}, peaks

    return {
        "id": "eeg-topography",
        "title": "Where each rhythm sits on the head",
        "licence": "ODC-BY 1.0",
        "citation": "PhysioNet EEG Motor Movement/Imagery Database, subject 1, "
                    "runs 1 and 2, all 64 channels. Electrode coordinates are "
                    "the standard 10-05 template (MNE), not this person's head.",
        "method": "Welch spectrum per electrode over 50 seconds; each band's "
                  "share of that electrode's own 1 to 45 Hz power, so a site "
                  "with a big signal cannot dominate every band.",
        "bands": [{"name": n, "low_Hz": lo, "high_Hz": hi} for n, lo, hi, _ in BANDS],
        "peaks": peaks,
        "channels": chans,
        "states": states,
    }


# ------------------------------------------------------------- myelin ----
def get_myelin():
    """Myelin water fraction against age, 45 people from 18 to 79.

    This is the one dataset here with a time axis measured in decades. It is
    also the only place on the scale map where the slow end is a measurement
    rather than a citation.
    """
    import openpyxl

    wb = openpyxl.load_workbook(cached(MYELIN_URL, "faizy2018_mwf.xlsx"),
                                data_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    ia = hdr.index("AGE")
    data = [r for r in rows[1:] if r and isinstance(r[ia], (int, float))]
    age = np.array([r[ia] for r in data], float)
    print(f"  {len(data)} subjects, ages {age.min():.0f} to {age.max():.0f}")

    # Two regions, chosen because they disagree: association white matter loses
    # myelin steadily across adult life, and the corticospinal tract, the tract
    # this site's other page draws running brainstem to motor cortex, does not.
    series = {}
    for col, label in [("parietal_WM_3mm", "Parietal white matter"),
                       ("CST_3mm", "Corticospinal tract")]:
        v = np.array([r[hdr.index(col)] for r in data], float)
        r = float(np.corrcoef(age, v)[0, 1])
        fit = np.polyfit(age, v, 1)
        order = np.argsort(age)
        series[col] = {
            "label": label, "n": len(v),
            "r_with_age": round(r, 3),
            "slope_per_decade": round(float(fit[0]) * 10, 5),
            "x": [round(float(a), 1) for a in age[order]],
            "y": [round(float(b), 4) for b in v[order]],
        }
        print(f"    {label:24s} r(age) = {r:+.3f}, "
              f"{fit[0]*10:+.5f} per decade")
    # The contrast is the point, so it is asserted rather than hoped for.
    assert series["parietal_WM_3mm"]["r_with_age"] < -0.7
    assert abs(series["CST_3mm"]["r_with_age"]) < 0.3
    return {
        "id": "myelin-lifespan",
        "title": "Myelin across adult life",
        "source": MYELIN_URL, "licence": "CC BY 4.0",
        "citation": "Faizy TD, Kumar D, Broocks G, et al. Age-related measures "
                    "of the human brain's white matter myelin. Sci Rep 8, 14991 "
                    "(2018). doi.org/10.1038/s41598-018-33112-8",
        "x_label": "years of age", "y_label": "myelin water fraction",
        "series": series,
    }


if __name__ == "__main__":
    print("action potential")
    ap = get_ap()
    print("eeg")
    eeg = get_eeg()
    print("seeg")
    seeg = get_seeg()
    print("the five rhythms")
    bands = get_bands()
    print("topography")
    topo = get_topography()
    print("myelin across the lifespan")
    mye = get_myelin()
    doc = {
        "about": "Real recordings and real measurements of real human brains. "
                 "Each carries its own source, licence and citation, because "
                 "they differ.",
        "signals": {s["id"]: s for s in (ap, eeg, bands, topo, seeg, mye)},
    }
    with open(p("data", "signals.json"), "w") as f:
        json.dump(doc, f, separators=(",", ":"))
    print(f"\nwrote data/signals.json "
          f"({os.path.getsize(p('data','signals.json'))/1e3:.0f} kB)")
